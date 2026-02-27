"""
Backup & Import utilities for the UA Manager system.
Exports all data as a single-sheet Excel file with labeled sections,
and imports from both single-sheet and multi-sheet formats.
"""
import io
from datetime import datetime, date
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q, Sum
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.parsers import MultiPartParser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers


# ── Model configs: (model_class, section_name, field_list) ──
def _get_model_configs():
    from .models import (
        Client, Contact, Invoice, Payment, Ticket,
        GeneticsSerial, SubscriptionModule, LivestockType,
        Todo, Project, Reminder, ActivityLog
    )
    return [
        (Contact, 'Contacts', [
            'id', 'client_id', 'name', 'phone', 'role',
        ]),
        (Invoice, 'Invoices', [
            'id', 'client_id', 'invoice_type', 'total_amount', 'status',
        ]),
        (Payment, 'Payments', [
            'id', 'invoice_id', 'amount', 'date', 'direction',
        ]),
        (Ticket, 'Tickets', [
            'id', 'client_id', 'issue_description', 'status', 'category', 'resolution_notes',
        ]),
        (GeneticsSerial, '4Genetics Serials', [
            'id', 'serial_number', 'client_id', 'product_type', 'is_active', 'assigned_date', 'notes',
        ]),
        (SubscriptionModule, 'Subscription Modules', [
            'id', 'name', 'description', 'is_active', 'order',
        ]),
        (LivestockType, 'Livestock Types', [
            'id', 'name', 'price_multiplier',
        ]),
        (Todo, 'Tasks', [
            'id', 'text', 'priority', 'status', 'due_date', 'is_done',
        ]),
        (Project, 'Projects', [
            'id', 'name', 'description', 'status',
        ]),
    ]


# ── Styling constants ──
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
SECTION_FONT = Font(name='Calibri', bold=True, size=13, color='1B5E20')
SECTION_FILL = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

# Client-specific premium styles
CLIENT_HEADER_FILL = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
CLIENT_SECTION_FILL = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
CLIENT_SECTION_FONT = Font(name='Calibri', bold=True, size=14, color='0D47A1')
CLIENT_EVEN_ROW = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
STATUS_ACTIVE_FONT = Font(name='Calibri', bold=True, color='2E7D32')
STATUS_EXPIRING_FONT = Font(name='Calibri', bold=True, color='E65100')
STATUS_EXPIRED_FONT = Font(name='Calibri', bold=True, color='C62828')

THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
EVEN_ROW_FILL = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')


def _format_cell_value(value):
    """Convert Python value to Excel-safe value."""
    if isinstance(value, (datetime, date)):
        return value.isoformat() if value else ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    return value


def _get_client_status(end_date):
    """Return status string based on subscription end date."""
    if not end_date:
        return 'Unknown'
    today = timezone.now().date()
    days_left = (end_date - today).days
    if days_left < 0:
        return 'Expired'
    if days_left <= 60:
        return 'Expiring Soon'
    return 'Active'


def _write_professional_clients_section(ws, current_row):
    """Write a premium, enriched Clients section with computed columns."""
    from .models import Client

    # Annotate clients with aggregated data
    clients = Client.objects.annotate(
        total_invoices=Count('invoices'),
        open_tickets=Count('tickets', filter=Q(tickets__status__in=['Open', 'In Progress'])),
        total_tickets=Count('tickets'),
        contacts_count=Count('contacts'),
        total_revenue=Sum('invoices__total_amount'),
        serials_count=Count('genetics_serials'),
    ).order_by('name')

    client_count = clients.count()

    # ── Section banner (premium blue) ──
    CLIENT_HEADERS = [
        'ID', 'Client Name', 'Farm Name', 'Phone', 'Serial Number',
        'Status', 'Subscription Start', 'Subscription End', 'Days Remaining',
        'Modules', 'Demo?', 'Demo Start', 'Demo End',
        'Contacts', 'Total Invoices', 'Total Revenue (EGP)',
        'Open Tickets', 'Total Tickets', '4G Serials',
        'Notes', 'Created', 'Last Updated',
    ]
    max_cols = len(CLIENT_HEADERS)

    # Banner row
    for col_idx in range(1, max_cols + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = CLIENT_SECTION_FILL
        cell.border = THIN_BORDER
    banner = ws.cell(row=current_row, column=1,
                     value=f'★  CLIENTS  —  {client_count} registered farms')
    banner.font = CLIENT_SECTION_FONT
    banner.fill = CLIENT_SECTION_FILL
    banner.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=max_cols)
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # ── Column headers ──
    for col_idx, header in enumerate(CLIENT_HEADERS, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = CLIENT_HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # ── Data rows ──
    today = timezone.now().date()
    if clients.exists():
        for idx, client in enumerate(clients):
            is_even = idx % 2 == 0
            days_remaining = (client.subscription_end_date - today).days if client.subscription_end_date else None
            status = _get_client_status(client.subscription_end_date)
            revenue = float(client.total_revenue) if client.total_revenue else 0

            row_data = [
                client.id,
                client.name,
                client.farm_name,
                client.phone,
                client.serial_number,
                status,
                client.subscription_start_date.isoformat() if client.subscription_start_date else '',
                client.subscription_end_date.isoformat() if client.subscription_end_date else '',
                days_remaining,
                client.subscription_modules or '',
                'Yes' if client.is_demo else 'No',
                client.demo_start_date.isoformat() if client.demo_start_date else '',
                client.demo_end_date.isoformat() if client.demo_end_date else '',
                client.contacts_count,
                client.total_invoices,
                revenue,
                client.open_tickets,
                client.total_tickets,
                client.serials_count,
                client.general_notes or '',
                client.created_at.strftime('%Y-%m-%d %H:%M') if client.created_at else '',
                client.updated_at.strftime('%Y-%m-%d %H:%M') if client.updated_at else '',
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if is_even:
                    cell.fill = CLIENT_EVEN_ROW

            # Status-specific font coloring (column 6)
            status_cell = ws.cell(row=current_row, column=6)
            if status == 'Active':
                status_cell.font = STATUS_ACTIVE_FONT
            elif status == 'Expiring Soon':
                status_cell.font = STATUS_EXPIRING_FONT
            elif status == 'Expired':
                status_cell.font = STATUS_EXPIRED_FONT

            # Days remaining: red if negative, orange if < 60
            if days_remaining is not None:
                days_cell = ws.cell(row=current_row, column=9)
                if days_remaining < 0:
                    days_cell.font = Font(bold=True, color='C62828')
                elif days_remaining <= 60:
                    days_cell.font = Font(bold=True, color='E65100')
                else:
                    days_cell.font = Font(color='2E7D32')

            # Revenue formatting
            rev_cell = ws.cell(row=current_row, column=16)
            rev_cell.number_format = '#,##0.00'

            current_row += 1
    else:
        empty_cell = ws.cell(row=current_row, column=1, value='(no clients registered)')
        empty_cell.font = Font(italic=True, color='9E9E9E')
        current_row += 1

    # Blank separator
    current_row += 1
    return current_row


class BackupExportView(APIView):
    """GET → Download full system backup as a styled single-sheet Excel file."""
    permission_classes = [AllowAny]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'UA Manager Backup'

        configs = _get_model_configs()
        current_row = 1

        # ── Title row ──
        title_cell = ws.cell(row=current_row, column=1, value='UA Manager — Full System Backup')
        title_cell.font = Font(name='Calibri', bold=True, size=16, color='1B5E20')
        current_row += 1
        ts_cell = ws.cell(row=current_row, column=1, value=f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        ts_cell.font = Font(name='Calibri', italic=True, size=10, color='757575')
        current_row += 2  # blank row

        # ── CLIENTS — Premium section ──
        current_row = _write_professional_clients_section(ws, current_row)

        # ── Other sections ──
        for Model, section_name, fields in configs:
            rows = list(Model.objects.all().values_list(*fields))
            headers = [f.replace('_', ' ').title() for f in fields]
            max_cols = len(headers)

            # Section banner
            for col_idx in range(1, max_cols + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = SECTION_FILL
                cell.border = THIN_BORDER
            banner = ws.cell(row=current_row, column=1,
                             value=f'▸ {section_name}  ({len(rows)} records)')
            banner.font = SECTION_FONT
            banner.fill = SECTION_FILL
            banner.alignment = Alignment(vertical='center')
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=max_cols)
            current_row += 1

            # Column headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
            current_row += 1

            # Data rows
            if rows:
                for data_idx, row_data in enumerate(rows):
                    is_even = data_idx % 2 == 0
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx,
                                       value=_format_cell_value(value))
                        cell.border = THIN_BORDER
                        if is_even:
                            cell.fill = EVEN_ROW_FILL
                    current_row += 1
            else:
                empty_cell = ws.cell(row=current_row, column=1, value='(no records)')
                empty_cell.font = Font(italic=True, color='9E9E9E')
                current_row += 1

            current_row += 1  # blank separator

        # ── Auto-width columns ──
        for col_idx in range(1, ws.max_column + 1):
            max_len = 8
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 45))
            ws.column_dimensions[col_letter].width = max_len + 3

        ws.freeze_panes = 'A5'

        # Write to response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
        filename = f'UA_Manager_Backup_{timestamp}.xlsx'

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class BackupImportView(APIView):
    """POST → Import data from an uploaded Excel backup file.
    Supports both single-sheet (section-based) and multi-sheet formats.
    """
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser]

    # Client field mapping for import (maps the enriched headers back to model fields)
    CLIENT_IMPORT_HEADERS = {
        'ID': 'id',
        'Client Name': 'name',
        'Farm Name': 'farm_name',
        'Phone': 'phone',
        'Serial Number': 'serial_number',
        'Modules': 'subscription_modules',
        'Demo?': 'is_demo',
        'Demo Start': 'demo_start_date',
        'Demo End': 'demo_end_date',
        'Subscription Start': 'subscription_start_date',
        'Subscription End': 'subscription_end_date',
        'Notes': 'general_notes',
    }

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            return Response({'error': f'Invalid Excel file: {str(e)}'}, status=400)

        configs = _get_model_configs()
        config_map = {name: (Model, fields) for Model, name, fields in configs}
        results = {}
        errors = []

        # Detect format: single-sheet or multi-sheet
        if len(wb.sheetnames) == 1 or 'UA Manager Backup' in wb.sheetnames:
            results, errors = self._import_single_sheet(ws=wb[wb.sheetnames[0]],
                                                        config_map=config_map)
        else:
            for sheet_name in wb.sheetnames:
                if sheet_name in ('Backup Summary',):
                    continue
                if sheet_name == 'Clients':
                    from .models import Client
                    section_result = self._import_clients_sheet(wb[sheet_name])
                    results['Clients'] = section_result
                    if section_result['errors']:
                        errors.extend([f'Clients - {e}' for e in section_result['errors']])
                    continue
                if sheet_name not in config_map:
                    continue
                Model, fields = config_map[sheet_name]
                ws = wb[sheet_name]
                section_result = self._import_sheet_data(ws, Model, fields, start_row=1)
                results[sheet_name] = section_result
                if section_result['errors']:
                    errors.extend([f'{sheet_name} - {e}' for e in section_result['errors']])

        return Response({
            'success': True,
            'results': results,
            'total_errors': len(errors),
            'errors': errors[:20],
        })

    def _import_single_sheet(self, ws, config_map):
        """Parse a single-sheet backup with section banners."""
        from .models import Client

        results = {}
        errors = []
        all_rows = list(ws.iter_rows(values_only=False))
        section_ranges = []

        for row_idx, row in enumerate(all_rows):
            cell_val = row[0].value
            if cell_val and isinstance(cell_val, str):
                # Match both '★  CLIENTS' and '▸ SectionName'
                if cell_val.startswith('★'):
                    section_ranges.append(('Clients', row_idx))
                elif cell_val.startswith('▸'):
                    section_name = cell_val.split('▸')[1].split('(')[0].strip()
                    section_ranges.append((section_name, row_idx))

        for i, (section_name, banner_row_idx) in enumerate(section_ranges):
            header_row_idx = banner_row_idx + 1
            if i + 1 < len(section_ranges):
                next_section_row = section_ranges[i + 1][1]
            else:
                next_section_row = len(all_rows)

            header_row = all_rows[header_row_idx]

            if section_name == 'Clients':
                # Use the client-specific header mapping
                col_map = {}
                for col_idx, cell in enumerate(header_row):
                    if cell.value and cell.value in self.CLIENT_IMPORT_HEADERS:
                        col_map[col_idx] = self.CLIENT_IMPORT_HEADERS[cell.value]

                created_count = 0
                updated_count = 0
                section_errors = []

                for row_idx in range(header_row_idx + 1, next_section_row):
                    row = all_rows[row_idx]
                    values = [cell.value for cell in row]
                    if not any(v is not None for v in values):
                        continue
                    if values[0] and isinstance(values[0], str) and values[0].startswith('(no'):
                        continue

                    data = {}
                    for col_idx, value in enumerate(values):
                        if col_idx in col_map:
                            field_name = col_map[col_idx]
                            if isinstance(value, str) and value.lower() in ('yes', 'no'):
                                value = value.lower() == 'yes'
                            if value == '' or value == 'None':
                                value = None
                            data[field_name] = value

                    if not data:
                        continue
                    try:
                        record_id = data.pop('id', None)
                        if record_id is not None:
                            try:
                                obj = Client.objects.get(pk=record_id)
                                for key, val in data.items():
                                    setattr(obj, key, val)
                                obj.save()
                                updated_count += 1
                            except Client.DoesNotExist:
                                Client.objects.create(id=record_id, **data)
                                created_count += 1
                        else:
                            Client.objects.create(**data)
                            created_count += 1
                    except Exception as e:
                        section_errors.append(f'Row {row_idx + 1}: {str(e)}')

                results['Clients'] = {'created': created_count, 'updated': updated_count, 'errors': section_errors}
                if section_errors:
                    errors.extend([f'Clients - {e}' for e in section_errors])

            elif section_name in config_map:
                Model, fields = config_map[section_name]
                label_to_field = {f.replace('_', ' ').title(): f for f in fields}
                col_map = {}
                for col_idx, cell in enumerate(header_row):
                    if cell.value and cell.value in label_to_field:
                        col_map[col_idx] = label_to_field[cell.value]

                created_count = 0
                updated_count = 0
                section_errors = []

                for row_idx in range(header_row_idx + 1, next_section_row):
                    row = all_rows[row_idx]
                    values = [cell.value for cell in row]
                    if not any(v is not None for v in values):
                        continue
                    if values[0] and isinstance(values[0], str) and values[0].startswith('(no'):
                        continue

                    data = {}
                    for col_idx, value in enumerate(values):
                        if col_idx in col_map:
                            field_name = col_map[col_idx]
                            if isinstance(value, str) and value.lower() in ('yes', 'no'):
                                value = value.lower() == 'yes'
                            if value == '' or value == 'None':
                                value = None
                            data[field_name] = value

                    if not data:
                        continue
                    try:
                        record_id = data.pop('id', None)
                        if record_id is not None:
                            try:
                                obj = Model.objects.get(pk=record_id)
                                for key, val in data.items():
                                    setattr(obj, key, val)
                                obj.save()
                                updated_count += 1
                            except Model.DoesNotExist:
                                Model.objects.create(id=record_id, **data)
                                created_count += 1
                        else:
                            Model.objects.create(**data)
                            created_count += 1
                    except Exception as e:
                        section_errors.append(f'Row {row_idx + 1}: {str(e)}')

                results[section_name] = {'created': created_count, 'updated': updated_count, 'errors': section_errors}
                if section_errors:
                    errors.extend([f'{section_name} - {e}' for e in section_errors])

        return results, errors

    def _import_clients_sheet(self, ws):
        """Import clients from a dedicated Clients sheet."""
        from .models import Client
        header_row = [cell.value for cell in ws[1]]
        col_map = {}
        for col_idx, header in enumerate(header_row):
            if header and header in self.CLIENT_IMPORT_HEADERS:
                col_map[col_idx] = self.CLIENT_IMPORT_HEADERS[header]

        created_count = 0
        updated_count = 0
        sheet_errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v is not None for v in row):
                continue
            data = {}
            for col_idx, value in enumerate(row):
                if col_idx in col_map:
                    field_name = col_map[col_idx]
                    if isinstance(value, str) and value.lower() in ('yes', 'no'):
                        value = value.lower() == 'yes'
                    if value == '' or value == 'None':
                        value = None
                    data[field_name] = value
            if not data:
                continue
            try:
                record_id = data.pop('id', None)
                if record_id is not None:
                    try:
                        obj = Client.objects.get(pk=record_id)
                        for key, val in data.items():
                            setattr(obj, key, val)
                        obj.save()
                        updated_count += 1
                    except Client.DoesNotExist:
                        Client.objects.create(id=record_id, **data)
                        created_count += 1
                else:
                    Client.objects.create(**data)
                    created_count += 1
            except Exception as e:
                sheet_errors.append(f'Row {row_idx}: {str(e)}')

        return {'created': created_count, 'updated': updated_count, 'errors': sheet_errors}

    def _import_sheet_data(self, ws, Model, fields, start_row=1):
        """Import data from a single sheet (multi-sheet format)."""
        header_row = [cell.value for cell in ws[start_row]]
        if not header_row or not any(header_row):
            return {'created': 0, 'updated': 0, 'errors': []}

        label_to_field = {f.replace('_', ' ').title(): f for f in fields}
        col_map = {}
        for col_idx, header in enumerate(header_row):
            if header in label_to_field:
                col_map[col_idx] = label_to_field[header]

        created_count = 0
        updated_count = 0
        sheet_errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row + 1, values_only=True), start=start_row + 1):
            if not any(v is not None for v in row):
                continue
            data = {}
            for col_idx, value in enumerate(row):
                if col_idx in col_map:
                    field_name = col_map[col_idx]
                    if isinstance(value, str) and value.lower() in ('yes', 'no'):
                        value = value.lower() == 'yes'
                    if value == '' or value == 'None':
                        value = None
                    data[field_name] = value
            if not data:
                continue
            try:
                record_id = data.pop('id', None)
                if record_id is not None:
                    try:
                        obj = Model.objects.get(pk=record_id)
                        for key, val in data.items():
                            setattr(obj, key, val)
                        obj.save()
                        updated_count += 1
                    except Model.DoesNotExist:
                        Model.objects.create(id=record_id, **data)
                        created_count += 1
                else:
                    Model.objects.create(**data)
                    created_count += 1
            except Exception as e:
                sheet_errors.append(f'Row {row_idx}: {str(e)}')

        return {'created': created_count, 'updated': updated_count, 'errors': sheet_errors}


class ImportPreviewView(APIView):
    """POST → Parse uploaded Excel and return sheet info + sample rows for mapping wizard."""
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            return Response({'error': f'Invalid Excel file: {str(e)}'}, status=400)

        # Build list of system fields (targets for mapping)
        configs = _get_model_configs()
        from .models import Client
        system_targets = {'Clients': [
            'id', 'name', 'farm_name', 'phone', 'serial_number',
            'subscription_modules', 'general_notes',
            'subscription_start_date', 'subscription_end_date',
            'is_demo', 'demo_start_date', 'demo_end_date',
        ]}
        for Model, section_name, fields in configs:
            system_targets[section_name] = fields

        # Parse each sheet
        sheets = []
        for sheet_name in wb.sheetnames:
            if sheet_name in ('Backup Summary',):
                continue
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True, max_row=min(ws.max_row, 6)))
            if not all_rows:
                continue

            # Detect header row (first row with text values)
            header_idx = 0
            for idx, row in enumerate(all_rows):
                if any(isinstance(v, str) and v.strip() for v in row if v is not None):
                    # Skip section banners (▸ or ★)
                    first_val = str(row[0]) if row[0] else ''
                    if first_val.startswith('▸') or first_val.startswith('★') or first_val.startswith('UA Manager'):
                        continue
                    header_idx = idx
                    break

            headers = [str(v) if v else f'Column {i+1}' for i, v in enumerate(all_rows[header_idx])] if header_idx < len(all_rows) else []
            # Sample rows (up to 3 after header)
            sample_rows = []
            for row in all_rows[header_idx + 1:header_idx + 4]:
                sample_rows.append([str(v) if v is not None else '' for v in row])

            sheets.append({
                'name': sheet_name,
                'columns': headers,
                'sample_rows': sample_rows,
                'row_count': ws.max_row - 1,
            })

        return Response({
            'sheets': sheets,
            'system_targets': system_targets,
        })


class SystemHealthView(APIView):
    """GET → Run system diagnostics and return health checks."""
    permission_classes = [AllowAny]

    def get(self, request):
        from .models import (
            Client, Contact, Invoice, Payment, Ticket,
            GeneticsSerial, SubscriptionModule, LivestockType,
            Todo, Project
        )

        checks = []
        today = timezone.now().date()

        # ── Client checks ──
        total_clients = Client.objects.count()
        active_clients = Client.objects.filter(subscription_end_date__gte=today).count()
        expired_clients = Client.objects.filter(subscription_end_date__lt=today).count()
        expiring_soon = Client.objects.filter(
            subscription_end_date__gte=today,
            subscription_end_date__lte=today + timezone.timedelta(days=60)
        ).count()
        no_phone = Client.objects.filter(phone='').count()

        checks.append({
            'category': 'Clients',
            'icon': '👥',
            'items': [
                {'label': 'Total clients', 'value': total_clients, 'status': 'ok' if total_clients > 0 else 'warning'},
                {'label': 'Active subscriptions', 'value': active_clients, 'status': 'ok'},
                {'label': 'Expired subscriptions', 'value': expired_clients, 'status': 'error' if expired_clients > 0 else 'ok',
                 'suggestion': f'Renew {expired_clients} expired subscriptions' if expired_clients > 0 else None},
                {'label': 'Expiring within 60 days', 'value': expiring_soon, 'status': 'warning' if expiring_soon > 0 else 'ok',
                 'suggestion': f'Follow up with {expiring_soon} clients about renewal' if expiring_soon > 0 else None},
                {'label': 'Missing phone numbers', 'value': no_phone, 'status': 'warning' if no_phone > 0 else 'ok',
                 'suggestion': f'Add phone numbers for {no_phone} clients' if no_phone > 0 else None},
            ]
        })

        # ── Invoice checks ──
        total_invoices = Invoice.objects.count()
        due_invoices = Invoice.objects.filter(status='Due').count()
        total_revenue = Invoice.objects.filter(status__in=['Paid to Us', 'Paid to Uniform']).aggregate(
            total=Sum('total_amount'))['total'] or 0
        total_due = Invoice.objects.filter(status='Due').aggregate(total=Sum('total_amount'))['total'] or 0

        checks.append({
            'category': 'Invoices & Revenue',
            'icon': '💰',
            'items': [
                {'label': 'Total invoices', 'value': total_invoices, 'status': 'ok'},
                {'label': 'Unpaid (Due)', 'value': due_invoices, 'status': 'error' if due_invoices > 3 else 'warning' if due_invoices > 0 else 'ok',
                 'suggestion': f'Collect {due_invoices} pending payments totaling {float(total_due):,.0f} EGP' if due_invoices > 0 else None},
                {'label': 'Total revenue collected', 'value': f'{float(total_revenue):,.0f} EGP', 'status': 'ok'},
                {'label': 'Outstanding amount', 'value': f'{float(total_due):,.0f} EGP', 'status': 'warning' if total_due > 0 else 'ok'},
            ]
        })

        # ── Tickets ──
        open_tickets = Ticket.objects.filter(status__in=['Open', 'In Progress']).count()
        total_tickets = Ticket.objects.count()
        stale_tickets = Ticket.objects.filter(
            status='Open',
            created_at__lte=timezone.now() - timezone.timedelta(days=7)
        ).count()

        checks.append({
            'category': 'Support Tickets',
            'icon': '🎫',
            'items': [
                {'label': 'Total tickets', 'value': total_tickets, 'status': 'ok'},
                {'label': 'Open / In Progress', 'value': open_tickets, 'status': 'warning' if open_tickets > 5 else 'ok'},
                {'label': 'Stale tickets (>7 days)', 'value': stale_tickets, 'status': 'error' if stale_tickets > 0 else 'ok',
                 'suggestion': f'Review {stale_tickets} stale tickets that need attention' if stale_tickets > 0 else None},
            ]
        })

        # ── Modules ──
        total_modules = SubscriptionModule.objects.count()
        active_modules = SubscriptionModule.objects.filter(is_active=True).count()
        unpriced = SubscriptionModule.objects.filter(price=0, is_active=True).count()

        checks.append({
            'category': 'Subscription Modules',
            'icon': '📦',
            'items': [
                {'label': 'Total modules', 'value': total_modules, 'status': 'ok' if total_modules > 0 else 'warning',
                 'suggestion': 'Add subscription modules to enable module-based invoicing' if total_modules == 0 else None},
                {'label': 'Active modules', 'value': active_modules, 'status': 'ok'},
                {'label': 'Modules without prices', 'value': unpriced, 'status': 'warning' if unpriced > 0 else 'ok',
                 'suggestion': f'Set prices for {unpriced} modules to enable auto-calculated invoices' if unpriced > 0 else None},
            ]
        })

        # ── Data quality ──
        orphan_contacts = Contact.objects.filter(client__isnull=True).count()
        clients_no_contacts = Client.objects.annotate(c=Count('contacts')).filter(c=0).count()
        serials = GeneticsSerial.objects.count()
        unassigned_serials = GeneticsSerial.objects.filter(client__isnull=True).count()

        checks.append({
            'category': 'Data Quality',
            'icon': '🔍',
            'items': [
                {'label': 'Clients without contacts', 'value': clients_no_contacts,
                 'status': 'warning' if clients_no_contacts > 0 else 'ok',
                 'suggestion': f'Add contact persons for {clients_no_contacts} clients' if clients_no_contacts > 0 else None},
                {'label': 'Orphaned contacts', 'value': orphan_contacts, 'status': 'error' if orphan_contacts > 0 else 'ok'},
                {'label': '4Genetics serials', 'value': serials, 'status': 'ok'},
                {'label': 'Unassigned serials', 'value': unassigned_serials,
                 'status': 'warning' if unassigned_serials > 0 else 'ok',
                 'suggestion': f'Assign {unassigned_serials} serials to clients' if unassigned_serials > 0 else None},
            ]
        })

        # Overall score
        total_items = sum(len(c['items']) for c in checks)
        ok_items = sum(1 for c in checks for i in c['items'] if i['status'] == 'ok')
        score = int((ok_items / total_items) * 100) if total_items > 0 else 100

        return Response({
            'score': score,
            'checks': checks,
            'summary': {
                'total_clients': total_clients,
                'active_clients': active_clients,
                'total_invoices': total_invoices,
                'open_tickets': open_tickets,
                'total_modules': total_modules,
            }
        })

